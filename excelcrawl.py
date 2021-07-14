# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook


class excelcrawl:
    def __init__(self):
        pass
    @staticmethod
    def search_subhotel_addr():
        load_wb = load_workbook("/Users/choi/Desktop/datamapping/ht_영국_matched_2nd_5th_3rd.xlsx")
        load_ws = load_wb['Sheet1']
        wb = openpyxl.Workbook()
        worksheet = wb.active
        for i in range(2,300):
            sub_hotel_addr = load_ws.cell(i,3).value
        wb.close()
        print(len(sub_hotel_addr))
        return sub_hotel_addr

    def search_hotels_name(self):
        load_wb = load_workbook("/Users/choi/Desktop/datamapping/ht_영국_matched_2nd_5th_3rd.xlsx")
        load_ws = load_wb['Sheet1']
        wb = openpyxl.Workbook()
        worksheet = wb.active
        for i in range(2,300):
            sub_hotel_addr = load_ws.cell(i,12).value
        wb.close()
        print(len(sub_hotel_addr))
        return sub_hotel_addr








    #위도,경도 값을 구글지도 에서 주소로 검색하여 reasponse받아서 기존 exel파일에 비어있는 위도 경도를 채움
    # 1차 위도 경도를 sub_hotel과 hotel을 비교해야함 (오차 범위내에서)
    # 1차 = 2차 범위내 통과하면 이름비교 맞으면 1 틀리면 2차진행
    # 2차 -1  위도 경도가 만약 틀리면 대문자를 소문자로 만들고 brute force를 사용 ( sub_hotel_addr과 hotel_addr 비교)
    # 2차 -2  만약 틀리면 대문자를 소문자로 만들고 kMP를 사용
    # 2차 -3  만약 틀리면 대문자를 소문자로 만들고 정규표현식을 사용
    #2차 =2 2차에서 맞으면 sub_hotel_cityname와 hotel_cityname을 비교한다 맞으면 2 틀리면 0
    #3차 틀리면 수기로 확인
    # 1이면 정답 2이면 애매 0 이면 틀림





    # print("1" + city_name, sub_city_name)





#print(google_map.location)

#lat = google_map.lat
#lng  = google_map.lng






#worksheet_write(1,1,lat)
#worksheet_write(1,2,lng)

